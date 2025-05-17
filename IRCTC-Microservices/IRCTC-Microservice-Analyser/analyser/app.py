# Import required modules

import os
import time
import http
import uuid
import json
import math
import typing
import fastapi
import uvicorn
import hashlib
import pydantic
import datetime
import openpyxl
import dateutil
import id_analyser
import bson.json_util
import irctc_analyser
import concurrent.futures
from mongo_connection import *
import fastapi_cache.decorator
import fastapi_cache.backends.inmemory 
from gchat_notification import notify_gchat

app = fastapi.FastAPI()

print(f'Reading env: {settings}')

class Job(pydantic.BaseModel):

    """Summary
    """

    uid: uuid.UUID = pydantic.Field(default_factory=uuid.uuid4)
    
    status: str = "in_progress"

    result: int = None

class FilesList(pydantic.BaseModel):
    """
    Base user model class for intercepting process request and reading variables from it
    """
    
    # Set the initial page number to '1' to display the first page of the results.
    page_id: typing.Optional[int] = 1

    # Set the file type of the uploaded file (AC, NON_AC, ARP, USER)
    # By default it is `USER`.
    ftype: typing.Optional[str] = 'USER'
    
    notuser: typing.Optional[bool] = True

class FileSearchDate(pydantic.BaseModel):

    # Set the date for whcih the analyzed file is to be searched.
    searchDate: typing.Optional[datetime.datetime] = pydantic.Field(default_factory=datetime.datetime.now)

    notuser: typing.Optional[bool] = True

jobs: typing.Dict[uuid.UUID, Job] = {}

def file_type(filename):
    '''
       Check the type of the file (AC, NON_aC, ARP) using the name of the file.

       :param filename: Name of the file to be analyzed.
       :ptype: str

       return: Type of the file.
       rtype: str
    '''

    # Check if the filename contains 'ARP' in it --> return file type as 'ARP'.
    if 'ARP' in filename: return 'ARP'
    
    # Check if the filename contains 'NO' in it --> return file type as 'NON_AC'.
    elif 'NO' in filename: return 'NON_AC'

    # Check if the filename contains 'AC' in it but not 'NO' in it --> return filetype as 'NON_AC'.
    elif ('AC' in filename and 'NO' not in filename): return 'AC'
    
    # Check if filename conatins 'cret' and 'id' in it --> return filetype as 'USER'.
    elif ('creat' in filename.lower() and 'id' in filename.lower()): return 'USER'

def convert_username_to_str(file_location):

    # open excel workbook
    workbook = openpyxl.load_workbook(file_location)

    # read by defualt first sheet which contains raw data
    sheet = workbook.worksheets[0]

    # column name to be checkedand converted to string
    column_name = "USERNAME"

    # find the index for USERNAME column
    column_mapping = {sheet[1][i].value: i for i in range(len(sheet[1]))}
    column_index = column_mapping.get(column_name)

    # iterate over the rows of USERNAME column
    for row_num in range(2, sheet.max_row + 1):
        # fetch the column value for that row
        column_value = sheet.cell(row=row_num, column=column_index+1).value  

        # check whether the USERNAME value is in datetime format or not
        if isinstance(column_value, (datetime.datetime, datetime.date)):

            # read the datetime format and convert to equivalent string value
            try:
                column_value_str = column_value.strftime('%b-%y')
            except:
                column_value_str = column_value.strftime('%b/%y')
        else:
            column_value_str = str(column_value)

        # save the cell value with the updated one
        sheet.cell(row=row_num, column=column_index+1).value = column_value_str  

    # close the workbook and save to same location
    workbook.save(file_location)
    workbook.close()

def get_file_size_in_MB(filepath):
    """
       Return the size of the file in MB.

       :param filepath: path of the file whose size is to be calculated.
       :ptype: str

       :return: size of the file in MB.
       :rtype: float
    """
    # Get the size of the file using the filepath.
    filesize = os.path.getsize(filepath)

    # Get the file size in MB.
    filesize_in_MB = filesize / 1024 / 1024

    # Round of the file size.
    filesize_in_MB = round(filesize_in_MB, 2)

    # Return the file size in MB.
    return filesize_in_MB

def file_in_mongo(filelocation, filename):
    """
       Update the status of the uploaded file in the database collection.

       :param1 filelocation: Location of the stored file.
       :param2 filename: Name of the file

       :p1type: str
       :p2type: str

       :return: Dictionary containing status of the file.
       :rtype: dict
    """
    
    # Open and read the file in binary mode.
    file = open(filelocation, 'rb')

    # Initialize the empty dictionary to store the status information of the file.
    file_mongo = dict()

    # Store the filename along the key `name`
    file_mongo['name'] = filename

    # Store the status of the file (Processing / Processed) along the key `status`.
    file_mongo['status'] = 'Processing'

    # Get the size of the file --> store the size along the key `filesize`.
    file_mongo['filesize'] = get_file_size_in_MB(filelocation)

    # Create a unique hash value for the file and store it along the key `hash`.
    file_mongo['hash'] = hashlib.md5(file.read()).hexdigest()

    # Stire the time at which the file is put under analysis along the key `ingest_date`.
    file_mongo['ingest_date'] = datetime.datetime.utcnow()

    # Get the type of the file (AC, NON_AC, ARP) --> store the filetype along the key `ftype`.
    file_mongo['ftype'] = file_type(filename)

    # Get the date of the the file from the file name --> store the date along the key `fdate`.
    file_mongo['fdate'] = dateutil.parser.parse(filename.split('_')[-1], fuzzy_with_tokens=True, dayfirst=True)[0]

    # Set the name of the analyzed file using its hash value.
    file_mongo['ana_f_name'] = file_mongo['hash'] + ".xlsx"

    # Return the dictionary conataining the file status.
    return file_mongo


@app.post('/file_ana', status_code=http.HTTPStatus.ACCEPTED)
def file_ana(file: fastapi.UploadFile, background_tasks: fastapi.BackgroundTasks):
    """
       Get the tatkal or userid file for the analysis.

       :param1 file:
       :param2 background_tasks: 

       :return1: Dictionary with information that more than 4 files can not be uploaded at a time.
       :return2: Location of the saved file.

       :rtype: dict
    """
    
    # Check if more than 4 files are under processing.
    if col_file_handle.count_documents({
        'status' : 'Processing'
    }) > 4:
        
        # Send notification to the google chat space.
        notify_gchat("Upload of More than 4 files attempted")

        # Return a response that maximum 4 files can be uploaded at a time.
        return {
            'status' : 500,
            'detail' : 'Maximum 4 files can be uploaded at a time.'
        }
    
    new_task = Job()

    jobs[new_task.uid] = new_task
    
    # Get the location of a file and store it in a variable.
    file_location = f"files/{file.filename}"

    # Open the file to write in a binary mode 
    with open(file_location, "wb+") as file_object:

        # Store the file in the `file_location`.
        file_object.write(file.file.read())

    # convert_username_to_str(file_location)

    # Get the status and information realated to the file
    file_rec = file_in_mongo(file_location, file.filename)

    # Chec if the file already does not exists in the database.
    if col_file_handle.count_documents({'hash': file_rec['hash']}, limit=1) >= 1:

        return {
            'status' : 403,
            'detail' : 'File with this name already exists'
        }

    # insert the file status and information into the the database.
    col_file_handle.insert_one(file_rec)

    # Send notification to the google chat space that the file is uploaded.
    notify_gchat(f"File uploaded : {file.filename}")
    print(file_rec)
    # Check if the uploaded file is a userid file
    if file_rec['ftype'] == 'USER':

        # Call IDAnalyser for file analysis --> add task as the background process.
        background_tasks.add_task(id_analyser.IDAnalyser, file_location, file_rec['hash'])
        
        # After file analysis return the message containing the detail fo the analyzed file location.
        return {
            'status' : 200,
            "detail": f"file '{file.filename}' saved at '{file_location}'"
        }
    
    # If the uploaded file is a PNR file.
    else:

        # Call IRCTCAnalyser for file analysis --> add task as the background process.
        background_tasks.add_task(irctc_analyser.IRCTCAnalyser, file_location, file_rec['ftype'], file_rec['hash'])

        # After file analysis return the message containing the detail fo the analyzed file location.
        return {
            'status' : 200,
            "detail": f"file '{file.filename}' saved at '{file_location}'"
        }


@app.get('/over-data')
def over_data():
    '''
    Get the data for the dashboard overview page.

    return: Dictionary containing the overview page data.
    '''

    print(f'Getting overview page data ..')

    # Get overview page data from the database --> convert to list.
    data = list(col_overdash_handle.find({},{'_id': 0}))

    # Return the data in the dictionary.
    return {
        'data': data
    }

@app.get("/log_download")
async def log_download(request: fastapi.Request):
    """
    Download the analysed file.

    :param request: Gets the request coming to the route.
    :ptype: fastapi.Request.

    :return: Return the file if found otherwise  a message that the file is not found.
    :rtype: dict
    """

    # Get data from the request made.
    data = await request.json()
    
    # Get the hash value of the file to be downloaded from the database.
    report_list = col_file_handle.find_one({
        "hash": data['f_hash']
    }, {'_id' : 0})

    # Get the hashed filename from the fetched records.
    log_file=report_list['ana_f_name']
    
    # Retur the analysed file response.
    return fastapi.responses.FileResponse(f'download_reports/{log_file}')

@fastapi_cache.decorator.cache(expire=60)
@app.post('/files_all', response_description="List of files embeded")
def list_files(file_list : FilesList):
    """
       Get all the file details and its analysis status and displays it on the dashboard.

       :param file_list: Get the data based on the FilesList model.
       :ptype: FilesList model.

       :return: Data to be displayed and count of pages.
       :rtype: dict
    """

    # Check if the details to be fetched are not for userid files.
    if file_list.notuser == True:

        # Get the details of the files whose 'ftype' is not `USER` or `SUS_USER` --> sort the data based on `fdate` --> get `10` records at a time --> convert the data to list.
        files_data = list(col_file_handle.find({
            '$and': [
                {
                    'ftype': {
                        '$ne': 'USER'
                    }
                }, {
                    'ftype': {
                        '$ne': 'SUS_USERS'
                    }
                }
            ]
        }, {}).sort('fdate',-1).skip((file_list.page_id - 1) * 10).limit(10))
        
        # Count the documents fetched where 'ftype' is not `USER` or `SUS_USER`.
        file_count = col_file_handle.count_documents({
            '$and': [
                {
                    'ftype': {
                        '$ne': 'USER'
                    }
                }, {
                    'ftype': {
                        '$ne': 'SUS_USERS'
                    }
                }
            ]
        })

    # Check if the details to be fetched are for userid files.
    else:

        # Get the details of the files whose `ftype` is `USER` or `SUS_USER` --> sort the data based on `fdate` --> get `10` records at a time --> convert the data to list.
        files_data = list(col_file_handle.find({
            'ftype': file_list.ftype
        }, {}).sort('fdate',-1).skip((file_list.page_id - 1) * 10).limit(10))

        # Count the documents fetched where 'ftype' is not `USER` or `SUS_USER`.
        file_count = col_file_handle.count_documents({
            'ftype': file_list.ftype
        })
    
    # Tracking list to store the filtered records.
    filtered_records_json = list()

    # Iterate over the fetched data.
    for data in files_data:

        # Convert each record into JSON string.
        rec = json.dumps(data, default=bson.json_util.default)

        # Add the JSON to the tracking list.
        filtered_records_json.insert(0, rec)
    
    # Return the dictionary containing filtered records and total page count of filtered records.
    page_count = math.ceil(file_count/10)

    # Return the data along with the total documents and the pages.
    return {
        'data_list': filtered_records_json[::-1], 
        'total_pages' : page_count,
        'total_rows': page_count
    }

@fastapi_cache.decorator.cache(expire=60)
@app.post('/files_find', response_description="List of files embeded")
def list_files_find(search_date: FileSearchDate):
    """
       Search file for the specific date.

       :param search_date: Date for which the analyzed are to be searched.
       :ptype: datetime
    """

    #
    new_date = datetime.datetime.combine(search_date.searchDate.date(), datetime.time.min)
    if search_date.notuser == True:
        files_data = list(col_file_handle.find({
            '$and': [
                {
                    'ftype': {
                        '$ne': 'USER'
                    }
                }, {
                    'ftype': {
                        '$ne': 'SUS_USERS'
                    }
                }
            ],
            'fdate': new_date
        }, {}))

        file_count = col_file_handle.count_documents({
            '$and': [
                {
                    'ftype': {
                        '$ne': 'USER'
                    }
                }, {
                    'ftype': {
                        '$ne': 'SUS_USERS'
                    }
                }
            ],
            'fdate': new_date
        })

    else:
        files_data = list(col_file_handle.find({
            'ftype': 'USER',
            'fdate': new_date
        }, {}))

        file_count = col_file_handle.count_documents({
            'ftype': 'USER',
            'fdate': new_date
        })

    filtered_records_json = []
    for data in files_data:
        rec = json.dumps(data, default=bson.json_util.default)
        filtered_records_json.insert(0, rec)
    
    # Return the dictionary containing filtered records and total page count of filtered records.
    page_count = math.ceil(file_count / 10)

    return {
        'data_list': filtered_records_json[::-1], 
        'total_pages' : page_count,
        'total_rows': page_count
    }

def setup_mount_folder(dir):
    """Summary

    Returns:
        TYPE: Description
    """
    path = os.getcwd()
    folder = os.path.join(path, dir)
    if not os.path.isdir(folder):
        os.mkdir(folder)
    return folder


async def startup_event():
    setup_mount_folder('files')
    setup_mount_folder('download_reports')
    app.state.executor = concurrent.futures.ProcessPoolExecutor()
    fastapi_cache.FastAPICache.init(fastapi_cache.backends.inmemory.InMemoryBackend(), prefix="fastapi-cache")

async def shutdown_event():
    app.state.executor.shutdown()

app.add_event_handler("startup", startup_event)
app.add_event_handler("shutdown", shutdown_event)

if __name__ == "__main__":
    uvicorn.run(
        'app:app', 
        host='0.0.0.0', 
        port=55562, 
        workers=4
    )